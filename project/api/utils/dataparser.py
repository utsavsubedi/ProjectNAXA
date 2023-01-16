from openpyxl import load_workbook
import json 
from django.conf import settings

class ExcelParser:
    def __init__(self, file_path):
        self.file_path = file_path 
        self.workbook = load_workbook(self.file_path, read_only=True)
        self.current_sheet = self.workbook.active
        self.columns = self.get_fields_name()
        self.field_values = None
        self.mapper_file = settings.BASE_DIR / 'data/model_excel_mapper.json'
        with open(self.mapper_file, 'r') as mapper:
            self.mapper = json.load(mapper)

    def get_fields_name(self):
        i =  1
        value_list = list()
        while self.current_sheet.cell(row=1, column=i).value!=None:
            value_list.append(self.current_sheet.cell(row=1, column=i).value.lower().replace(' ', '_'))
            i += 1
        return value_list

    def get_fields_value_dict(self):
        columns = self.columns
        self.field_values = dict()
        for i in range(len(columns)):
            current_col = self.mapper[columns[i]]
            self.field_values[current_col] = list()
            j = 1
            while self.current_sheet.cell(row=j+1, column=i+1).value != None:
                current_value = self.current_sheet.cell(row=j+1, column=i+1).value
                self.field_values[current_col].append(current_value)
                j += 1
        return self.validate_data()

    def validate_data(self):
        max_values = list()
        for column in self.columns:
            max_values.append(len(self.field_values[self.mapper[column]]))
        if max(max_values) != min(max_values):
            return None
        return self.field_values